"""Excel (.xlsx) intake for engagement leads: downloadable template and importer.

Explicitly Excel, never CSV. The import is defensive: the upload is size-bounded and
must really be an .xlsx workbook (a malformed or hostile file is rejected, never
parsed loosely), and every row goes through the same rules and de-duplication as the
public form and manual entry.
"""
# ruff: noqa: E501 - long lines kept readable
from __future__ import annotations

import io
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from fastapi.responses import Response

from . import audit
from .engagement import EngagementIn, enregistrer_invitation
from .permissions_rbac import require_permission
from .schemas import UserMe

router = APIRouter(prefix="/api/v1/admin/engagement", tags=["engagement"])

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_MAX_BYTES = 2 * 1024 * 1024  # 2 MB is ample for a lead list
_COLONNES = ["email", "prenoms", "nom", "telephone", "indicatif_pays", "code_pays"]


@router.get("/template.xlsx")
def template_xlsx(user: Annotated[UserMe, Depends(require_permission("inscriptions.gerer"))]) -> Response:
    """Download the exact .xlsx template to fill for the engagement import."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "engagements"
    ws.append(_COLONNES)
    ws.append(["marie.koffi@example.com", "Marie", "KOFFI", "0700000000", "+225", "CI"])
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": 'attachment; filename="modele-engagements.xlsx"'},
    )


def _cellule(row: tuple, idx: int) -> str | None:
    if idx >= len(row) or row[idx] is None:
        return None
    val = str(row[idx]).strip()
    return val or None


@router.post("/import")
async def importer_xlsx(
    fichier: UploadFile,
    user: Annotated[UserMe, Depends(require_permission("inscriptions.gerer"))],
) -> dict[str, object]:
    """Import engagement leads from an .xlsx file (columns of the template).

    Duplicate e-mails (in the file or already captured) are skipped and reported;
    one bad row never aborts the import. Non-xlsx or oversized files are refused."""
    from openpyxl import load_workbook

    content = await fichier.read()
    if len(content) > _MAX_BYTES:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="fichier trop volumineux (max 2 Mo)")
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="fichier vide")
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - any parse error means it is not a valid xlsx
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="fichier Excel (.xlsx) invalide ou corrompu") from exc
    ws = wb.active

    crees = 0
    doublons: list[str] = []
    erreurs: list[dict[str, str]] = []
    vus: set[str] = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header row
        email = _cellule(row, 0)
        if not email:
            continue  # skip blank lines
        low = email.lower()
        if low in vus:
            doublons.append(low)
            continue
        vus.add(low)
        try:
            payload = EngagementIn(
                email=email,
                prenoms=_cellule(row, 1),
                nom=_cellule(row, 2),
                telephone=_cellule(row, 3),
                pays_indicatif=_cellule(row, 4),
                pays_code=_cellule(row, 5),
            )
        except Exception:  # noqa: BLE001 - validation error on a row, report and continue
            erreurs.append({"email": low, "raison": "ligne invalide (e-mail ou champ incorrect)"})
            continue
        try:
            enregistrer_invitation(payload, source="import", cree_par=user.id, role=user.role)
            crees += 1
        except HTTPException as exc:
            if exc.status_code == 409:
                doublons.append(low)
            else:
                erreurs.append({"email": low, "raison": str(exc.detail)})
    audit.log(user.id, user.role, "import_engagement", "invitation_engagement", None, {"crees": crees, "doublons": len(doublons), "erreurs": len(erreurs)})
    return {"crees": crees, "doublons": doublons, "erreurs": erreurs}
